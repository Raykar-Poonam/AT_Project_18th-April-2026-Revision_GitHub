import openpyxl


def row_num(filename, sheet):
    Excel_File = openpyxl.load_workbook(filename)
    Sheet = Excel_File[sheet]
    return Sheet.max_row


def readData(filename, sheet, row, column):
    Excel_File = openpyxl.load_workbook(filename)
    Sheet = Excel_File[sheet]
    return Sheet.cell(row=row, column=column).value


def writeData(filename, sheet, row, column, data):
    Excel_File = openpyxl.load_workbook(filename)
    Sheet = Excel_File[sheet]
    Sheet.cell(row=row, column=column).value = data
    Excel_File.save(filename)
